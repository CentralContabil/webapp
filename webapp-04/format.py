from __future__ import annotations
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import unicodedata
import re

# ==============================
# Configurações visuais gerais
# ==============================
HEADER_FONT   = Font(bold=True)
HEADER_HEIGHT = 28   # padrão global; Confronto sobrescreve para 45
ROW_HEIGHT    = 22

# Cores (ARGB)
HEADER_FILL = PatternFill(start_color="FFC5D9F1", end_color="FFC5D9F1", fill_type="solid")  # C5D9F1
ZEBRA_FILL  = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")  # DCE6F1

# ==============================
# Funções auxiliares
# ==============================
def _auto_width(ws: Worksheet) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[letter].width = min(max(8, max_len + 2), 60)

def _fixed_width(ws: Worksheet, width: float = 12.5) -> None:
    """Define largura fixa para todas as colunas existentes."""
    ws.sheet_format.defaultColWidth = width
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def _center_all(ws: Worksheet) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str):
                cell.value = cell.value.replace("\\n", " ").replace("\\r", " ")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

def _header_and_row_heights(ws: Worksheet) -> None:
    if ws.max_row >= 1:
        ws.row_dimensions[1].height = HEADER_HEIGHT
        for c in ws[1]:
            if isinstance(c, MergedCell):
                continue
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
    for r in range(2, (ws.max_row or 1) + 1):
        ws.row_dimensions[r].height = ROW_HEIGHT

def _freeze_header(ws: Worksheet) -> None:
    ws.freeze_panes = "A2"  # congela linha 1

def _highlight_delta_headers(ws: Worksheet) -> None:
    """Destaca cabeçalhos que começam com 'Delta' (case-sensitive)."""
    fill = PatternFill(start_color="FFFDE9D9", end_color="FFFDE9D9", fill_type="solid")
    if ws.max_row >= 1:
        for cell in ws[1]:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.strip().startswith("Delta"):
                cell.fill = fill

def _borders_and_stripes(ws: Worksheet) -> None:
    """Borda fina + zebra nas linhas pares (exceto cabeçalho)."""
    thin   = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    max_r, max_c = ws.max_row, ws.max_column
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c), start=1):
        is_zebra = (r_idx > 1 and r_idx % 2 == 0)
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.border = border
            if is_zebra:
                cell.fill = ZEBRA_FILL

def _borders_no_stripes(ws: Worksheet) -> None:
    """Apenas bordas finas; não pinta linhas (preserva cabeçalho/divergências)."""
    thin   = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.border = border

def _space_items_xml(ws: Worksheet) -> None:
    """Insere 2 linhas em branco quando a CHAVE mudar (para 'Itens (XML)')."""
    if ws.max_row < 3:
        return
    header = [c.value for c in ws[1]]
    try:
        chave_col = header.index("CHAVE") + 1  # 1-based
    except ValueError:
        return
    r = ws.max_row
    while r > 2:
        atual = ws.cell(row=r,   column=chave_col).value
        acima = ws.cell(row=r-1, column=chave_col).value
        if atual and acima and atual != acima:
            ws.insert_rows(r, amount=2)
        r -= 1





# -----------------------------
# Coerção numérica por cabeçalho
# -----------------------------
def _norm_header(txt: str) -> str:
    """Remove acentos, deixa UPPER, tira sufixos ' (SCI)'/' (Cliente)' para comparar."""
    if txt is None:
        return ""
    s = str(txt)
    # remove sufixo entre parênteses: "Vlr contábil (SCI)" -> "Vlr contábil"
    if " (" in s:
        s = s.split(" (", 1)[0]
    # normaliza acentos e caixa
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    return s.upper().strip()

_NUMERIC_FLOAT_HEADERS = {
    "VLR CONTABIL", "VALOR CONTABIL", "VLR CONTÁBIL", "VALOR CONTÁBIL",
    "BASE DE ICMS", "BASE ICMS",
    "VALOR DO ICMS", "VALOR ICMS",
    "BASE DE IPI", "BASE IPI",
    "VALOR DO IPI", "VALOR IPI",
    # variações com "SOMA DE ..."
    "SOMA DE VLR. CONTABIL", "SOMA DE VLR. CONTÁBIL",
    "SOMA DE BASE ICMS",
    "SOMA DE VLR. ICMS",
    "SOMA DE BASE IPI",
    "SOMA DE VLR. IPI",
}

_NUMERIC_INT_HEADERS = {"NOTA", "Nº NF.", "Nº NF", "NO NF", "NUM NF", "NUMERO NF", "NF"}

def _parse_float_like(v):
    """Converte strings com R$, ponto de milhar e vírgula decimal em float."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    # remove moeda e espaços
    s = s.replace("R$", "").replace("\u00A0", " ").replace(" ", "")
    # troca separadores: 1.234.567,89 -> 1234567.89
    s = s.replace(".", "").replace(",", ".")
    # mantém apenas dígitos, ponto e sinal
    s = re.sub(r"[^0-9\.\-]", "", s)
    try:
        return float(s) if s not in ("", "-", ".", "-.") else None
    except Exception:
        return None

def _parse_int_like(v):
    """Extrai dígitos de algo tipo '001234' e devolve int (ou None se não houver)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return int(round(float(v)))
        except Exception:
            return None
    m = re.search(r"(\d+)", str(v))
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None

def _coerce_numeric_by_header(ws: Worksheet) -> None:
    """Converte texto->número nas colunas reconhecidas e aplica number_format."""
    if ws.max_row < 2 or ws.max_column < 1:
        return

    # mapeia colunas-alvo pelo cabeçalho
    header_vals = [c.value for c in ws[1]]
    targets = {}  # col_idx -> ("int"|"float")
    for idx, h in enumerate(header_vals, start=1):
        nh = _norm_header(h)
        if nh in _NUMERIC_FLOAT_HEADERS:
            targets[idx] = "float"
        elif nh in _NUMERIC_INT_HEADERS:
            targets[idx] = "int"

    if not targets:
        return

    # percorre linhas e converte valor + aplica formato
    for r in range(2, ws.max_row + 1):
        for c_idx, kind in targets.items():
            cell = ws.cell(row=r, column=c_idx)
            # ignora se já for número
            if isinstance(cell.value, (int, float)):
                # só garante o formato
                if kind == "float":
                    cell.number_format = "#,##0.00"
                else:
                    cell.number_format = "0"
                continue

            if kind == "float":
                val = _parse_float_like(cell.value)
                if val is not None:
                    cell.value = float(val)
                    cell.number_format = "#,##0.00"
            else:  # int
                val = _parse_int_like(cell.value)
                if val is not None:
                    cell.value = int(val)
                    cell.number_format = "0"

# --------------------------
# Confronto: recursos extras
# --------------------------
def _insert_vertical_separator(ws: Worksheet, col_idx: int = 8, sep_width: float = 3.5) -> int:
    """
    Insere uma coluna vazia em col_idx (ex.: 8 = H),
    mescla de H1 até H(max_row) e mantém sem borda/cor.
    Retorna o índice da coluna separadora.
    """
    ws.insert_cols(col_idx, 1)
    max_r = max(ws.max_row, 1)
    ws.merge_cells(start_row=1, start_column=col_idx, end_row=max_r, end_column=col_idx)

    mc = ws.cell(row=1, column=col_idx)  # célula top-left do merge
    mc.value = None
    mc.fill = PatternFill()
    mc.border = Border()
    mc.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(col_idx)].width = sep_width
    return col_idx

def _tune_confronto_header(ws: Worksheet, height: float = 45) -> None:
    """Cabeçalho da Confronto: altura 45 + wrap."""
    ws.row_dimensions[1].height = height
    for cell in ws[1]:
        if isinstance(cell, MergedCell):
            continue
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _highlight_differences(ws: Worksheet, sep_col_idx: int | None = None) -> None:
    """
    Pinta toda a linha em aqua quando houver diferença
    e deixa as células divergentes em vermelho.
    """
    aqua_fill = PatternFill(start_color="FFDAEEF3", end_color="FFDAEEF3", fill_type="solid")
    red_font  = Font(color="FFFF0000")
    if ws.max_row < 2:
        return

    header = [c.value for c in ws[1]]
    pares = [
        ("Vlr contábil (SCI)", "Vlr contábil (Cliente)"),
        ("Base de ICMS (SCI)", "Base de ICMS (Cliente)"),
        ("Valor do ICMS (SCI)", "Valor do ICMS (Cliente)"),
        ("Base de IPI (SCI)", "Base de IPI (Cliente)"),
        ("Valor do IPI (SCI)", "Valor do IPI (Cliente)"),
    ]

    for r in range(2, ws.max_row + 1):
        linha_diff = False
        for sci_col, cli_col in pares:
            try:
                c_sci = header.index(sci_col) + 1
                c_cli = header.index(cli_col) + 1
            except ValueError:
                continue

            v_sci = ws.cell(row=r, column=c_sci).value or 0
            v_cli = ws.cell(row=r, column=c_cli).value or 0
            try:
                if abs(float(v_sci) - float(v_cli)) > 0.005:
                    ws.cell(row=r, column=c_sci).font = red_font
                    ws.cell(row=r, column=c_cli).font = red_font
                    linha_diff = True
            except Exception:
                pass

        if linha_diff:
            for c_idx in range(1, ws.max_column + 1):
                if sep_col_idx is not None and c_idx == sep_col_idx:
                    continue
                cell = ws.cell(row=r, column=c_idx)
                if not isinstance(cell, MergedCell):
                    cell.fill = aqua_fill

# ==============================
# Orquestração principal
# ==============================
def format_workbook(wb: Workbook) -> None:
    """Aplica formatação padrão em todas as abas + regras específicas."""
    for ws in wb.worksheets:
        name = (ws.title or "").strip().lower()

        # Base comum
        _header_and_row_heights(ws)
        _coerce_numeric_by_header(ws)   # <--- NOVO (garante números + number_format)
        _freeze_header(ws)
        _center_all(ws)
        _highlight_delta_headers(ws)

        if name == "confronto":
            # Regras específicas Confronto
            _fixed_width(ws, 12.5)                # largura fixa
            sep_col = _insert_vertical_separator(ws, col_idx=7, sep_width=12.5)
            ws.sheet_view.showGridLines = False   # remove linhas de grade
            _borders_no_stripes(ws)               # sem listras; preserva linhas em branco
            ws.cell(1, sep_col).border = Border() # separador sem borda
            _tune_confronto_header(ws, height=45) # cabeçalho 45 + wrap
            _highlight_differences(ws, sep_col_idx=sep_col)
        else:
            # Demais abas mantêm auto width + listras
            _auto_width(ws)
            _borders_and_stripes(ws)

    # Espaçamento específico para a aba de itens XML (se existir)
    if "Itens (XML)" in wb.sheetnames:
        _space_items_xml(wb["Itens (XML)"])
