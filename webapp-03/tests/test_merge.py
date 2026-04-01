# -*- coding: utf-8 -*-
from __future__ import annotations

import subprocess
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))
_FIXTURE_SPED = _ROOT.parent / "webapp-01" / "tests" / "fixtures" / "sped_minimo.txt"
_ENGINE = _ROOT.parent / "webapp-02" / "sped_engine"
_CLI_EXPORT = _ENGINE / "cli.py"
_CLI_MERGE = _ROOT / "cli_merge.py"
from merger import merge_sped_from_xlsx


@pytest.mark.skipif(not _FIXTURE_SPED.is_file(), reason="fixture SPED ausente")
@pytest.mark.skipif(not _CLI_EXPORT.is_file(), reason="sped_engine ausente")
def test_roundtrip_minimo_unchanged(tmp_path: Path) -> None:
    xlsx = tmp_path / "out.xlsx"
    out_txt = tmp_path / "merged.txt"
    subprocess.run(
        [sys.executable, str(_CLI_EXPORT), "--input", str(_FIXTURE_SPED), "--output", str(xlsx)],
        cwd=str(_ENGINE),
        check=True,
        capture_output=True,
        text=True,
    )
    subprocess.run(
        [sys.executable, str(_CLI_MERGE), "--sped", str(_FIXTURE_SPED), "--xlsx", str(xlsx), "--output", str(out_txt)],
        cwd=str(_ROOT),
        check=True,
        capture_output=True,
        text=True,
    )
    orig = _FIXTURE_SPED.read_text(encoding="utf-8", errors="replace")
    merged = out_txt.read_text(encoding="utf-8", errors="replace")
    assert merged.rstrip("\r\n") == orig.rstrip("\r\n")


@pytest.mark.skipif(not _FIXTURE_SPED.is_file(), reason="fixture SPED ausente")
@pytest.mark.skipif(not _CLI_EXPORT.is_file(), reason="sped_engine ausente")
def test_merge_aplica_edicao_reg_0000(tmp_path: Path) -> None:
    xlsx = tmp_path / "out_0000.xlsx"
    out_txt = tmp_path / "merged_0000.txt"
    subprocess.run(
        [sys.executable, str(_CLI_EXPORT), "--input", str(_FIXTURE_SPED), "--output", str(xlsx), "--sheets", "0000"],
        cwd=str(_ENGINE),
        check=True,
        capture_output=True,
        text=True,
    )

    wb = load_workbook(xlsx)
    ws = wb["0000"]
    headers = [str(c.value or "") for c in ws[1]]
    col_nome = headers.index("NOME") + 1
    ws.cell(row=2, column=col_nome).value = "EMPRESA EDITADA TESTE"
    wb.save(xlsx)

    subprocess.run(
        [sys.executable, str(_CLI_MERGE), "--sped", str(_FIXTURE_SPED), "--xlsx", str(xlsx), "--output", str(out_txt)],
        cwd=str(_ROOT),
        check=True,
        capture_output=True,
        text=True,
    )

    merged = out_txt.read_text(encoding="utf-8", errors="replace")
    assert "EMPRESA EDITADA TESTE" in merged


def test_merge_aplica_reg_generico_fora_dos_headers(tmp_path: Path) -> None:
    sped = tmp_path / "orig.txt"
    xlsx = tmp_path / "edit.xlsx"
    out = tmp_path / "out.txt"
    sped.write_text("|0000|017|0|\n|Z999|A|B|\n", encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "Z999"
    ws.append(["_LINHA", "COL_01", "COL_02", "COL_03"])
    ws.append([2, "Z999", "EDITADO", "B"])
    wb.save(xlsx)

    merge_sped_from_xlsx(sped, xlsx, out)

    merged = out.read_text(encoding="utf-8", errors="replace")
    assert "|Z999|EDITADO|B|" in merged


def test_merge_generico_usa_template_original_para_data_e_numero(tmp_path: Path) -> None:
    sped = tmp_path / "orig_template.txt"
    xlsx = tmp_path / "edit_template.xlsx"
    out = tmp_path / "out_template.txt"
    sped.write_text("|0000|017|0|\n|Z999|01022026|16664,42|ABC|\n", encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "Z999"
    ws.append(["_LINHA", "COL_01", "COL_02", "COL_03", "COL_04"])
    ws.append([2, "Z999", "2026-02-02 00:00:00", "16.664,42", "ABC"])
    wb.save(xlsx)

    merge_sped_from_xlsx(sped, xlsx, out)

    merged = out.read_text(encoding="utf-8", errors="replace")
    assert "|Z999|02022026|16664,42|ABC|" in merged
    assert "2026-02-02" not in merged
    assert "16664.42" not in merged


def test_merge_preserva_acentuacao_cp1252_no_sped_original(tmp_path: Path) -> None:
    sped = tmp_path / "orig_cp1252.txt"
    xlsx = tmp_path / "edit_cp1252.xlsx"
    out = tmp_path / "out_cp1252.txt"

    sped.write_bytes("|0500|01012026|A|TESTE|Despesa com Veículos|\n".encode("cp1252"))

    wb = Workbook()
    ws = wb.active
    ws.title = "0500"
    ws.append(["_LINHA", "COL_01", "COL_02", "COL_03", "COL_04", "COL_05"])
    ws.append([1, "0500", "01012026", "A", "TESTE", "Despesa com Veículos"])
    wb.save(xlsx)

    merge_sped_from_xlsx(sped, xlsx, out)

    merged = out.read_bytes()
    assert b"Ve\xedculos" in merged  # 'í' em cp1252
    assert b"\xef\xbf\xbd" not in merged  # caractere de substituição UTF-8 (�)


def test_merge_preserva_quantidade_de_campos_por_linha_no_generico(tmp_path: Path) -> None:
    sped = tmp_path / "orig_fields.txt"
    xlsx = tmp_path / "edit_fields.xlsx"
    out = tmp_path / "out_fields.txt"
    sped.write_text(
        "|K010||\n"
        "|1010|N|N|N|N|N|N|N|N|N|N|N|N|N|\n",
        encoding="utf-8",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "K010"
    ws.append(["_LINHA", "COL_01", "COL_02", "COL_03"])
    ws.append([1, "K010", "", ""])
    wb2 = wb.create_sheet("1010")
    wb2.append(
        ["_LINHA", "COL_01", "COL_02", "COL_03", "COL_04", "COL_05", "COL_06", "COL_07", "COL_08", "COL_09", "COL_10", "COL_11", "COL_12", "COL_13", "COL_14", "COL_15", "COL_16"]
    )
    wb2.append([2, "1010", "N", "N", "N", "N", "N", "N", "N", "N", "N", "N", "N", "N", "N", "", ""])
    wb.save(xlsx)

    merge_sped_from_xlsx(sped, xlsx, out)

    lines = out.read_text(encoding="utf-8", errors="replace").splitlines()
    k010_inner = lines[0].split("|")[1:-1]
    r1010_inner = lines[1].split("|")[1:-1]
    assert len(k010_inner) == 2
    assert len(r1010_inner) == 14


def test_merge_sem_sped_original_gera_saida_por_linha(tmp_path: Path) -> None:
    xlsx = tmp_path / "only.xlsx"
    out = tmp_path / "out_only.txt"
    wb = Workbook()
    ws = wb.active
    ws.title = "C100"
    ws.append(["_LINHA", "REG", "IND_OPER", "IND_EMIT", "COD_PART", "COD_MOD", "COD_SIT", "SER", "NUM_DOC", "CHV_NFE", "DT_DOC", "DT_E_S", "VL_DOC"])
    ws.append([1, "C100", "1", "0", "P", "55", "00", "1", "123", "CHV", "02022026", "02022026", "100,00"])
    wb.save(xlsx)

    merge_sped_from_xlsx(None, xlsx, out)

    lines = out.read_text(encoding="utf-8", errors="replace").splitlines()
    assert len(lines) == 1
    assert lines[0].startswith("|C100|1|0|P|55|00|1|123|CHV|02022026|02022026|100,00|")


def test_merge_sem_sped_original_ajusta_k010_e_remove_cauda_vazia(tmp_path: Path) -> None:
    xlsx = tmp_path / "only_k.xlsx"
    out = tmp_path / "out_k.txt"
    wb = Workbook()
    ws = wb.active
    ws.title = "K010"
    ws.append(["_LINHA", "REG", "DT_INI", "DT_FIN"])
    ws.append([1, "K010", "", ""])
    ws2 = wb.create_sheet("1010")
    ws2.append(
        [
            "_LINHA",
            "REG",
            "IND_EXP",
            "IND_CCRF",
            "IND_COMB",
            "IND_USINA",
            "IND_VA",
            "IND_EE",
            "IND_CART",
            "IND_FORM",
            "IND_AER",
            "IND_GIAF1",
            "IND_GIAF2",
            "IND_GIAF3",
            "IND_GIAF4",
            "IND_REST_RESSARC_COMPL",
            "IND_SIMPL",
            "IND_RTR",
        ]
    )
    ws2.append([2, "1010", "N", "N", "N", "N", "N", "N", "N", "N", "N", "N", "N", "N", "N", "", "", ""])
    wb.save(xlsx)

    merge_sped_from_xlsx(None, xlsx, out)

    lines = out.read_text(encoding="utf-8", errors="replace").splitlines()
    assert lines[0] == "|K010|0|"
    assert lines[1] == "|1010|N|N|N|N|N|N|N|N|N|N|N|N|N|"


def test_merge_sem_sped_original_sanitiza_cedilha_e_til(tmp_path: Path) -> None:
    xlsx = tmp_path / "only_0500.xlsx"
    out = tmp_path / "out_0500.txt"
    wb = Workbook()
    ws = wb.active
    ws.title = "0500"
    ws.append(["_LINHA", "REG", "DT_ALT", "COD_NAT_CC", "IND_CTA", "NIVEL", "NOME_CTA"])
    ws.append([1, "0500", "01012026", "06", "A", "5", "Ação ~ São João"])
    wb.save(xlsx)

    merge_sped_from_xlsx(None, xlsx, out)

    line = out.read_text(encoding="utf-8", errors="replace").splitlines()[0]
    assert "Ação" not in line
    assert "São" not in line
    assert "~" not in line
    assert "Acao  Sao Joao" in line


def test_inspect_xlsx_exige_sped_original_quando_layout_incompleto(tmp_path: Path) -> None:
    xlsx = tmp_path / "only_tail.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "K010"
    ws.append(["_LINHA", "REG", "COL_02", "COL_03"])
    ws.append([1, "K010", "0", ""])
    ws2 = wb.create_sheet("1010")
    ws2.append(["_LINHA", "REG", "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "EXTRA_01"])
    ws2.append([2, "1010", "N", "N", "N", "N", "N", "N", "N", "N", "N", "N", "N", "N", "N", "", ""])
    wb.save(xlsx)
    from inspect_xlsx import inspect_xlsx
    inspected = inspect_xlsx(xlsx)
    assert inspected["requiresOriginal"] is True
    assert any("colunas obrigatórias ausentes" in r for r in inspected["reasons"])


def test_merge_normaliza_data_e_valor_no_c100(tmp_path: Path) -> None:
    sped = tmp_path / "orig_c100.txt"
    xlsx = tmp_path / "edit_c100.xlsx"
    out = tmp_path / "out_c100.txt"
    sped.write_text(
        "|0000|017|0|\n"
        "|C100|1|0|PART|55|00|1|1|CHAVE|01022026|01022026|100|1|0|0|100|1|0|0|0|100|18|0|0|0|0|0|0|0|\n",
        encoding="utf-8",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "C100"
    ws.append(["_LINHA", "REG", "DT_DOC", "DT_E_S", "VL_DOC", "VL_MERC"])
    ws.append([2, "C100", datetime(2026, 2, 2, 0, 0, 0), "2026-02-02 00:00:00", "16.664,42", "16.556,80"])
    wb.save(xlsx)

    merge_sped_from_xlsx(sped, xlsx, out)

    merged = out.read_text(encoding="utf-8", errors="replace")
    assert "|02022026|02022026|16664,42|" in merged
    assert "2026-02-02" not in merged
    assert "16.664,42" not in merged
    assert "16664.42" not in merged


def test_inner_payload_c170_skips_injected() -> None:
    sys.path.insert(0, str(_ROOT.parent / "webapp-02" / "sped_engine"))
    from config import HEADERS  # noqa: WPS433

    from line_builders import build_sped_line, inner_payload_for_register

    row = {
        "REG": "C170",
        "NUM_DOC": "999",
        "CHV_NFE": "x",
        "NUM_ITEM": "1",
        "COD_ITEM": "ABC",
    }
    inner = inner_payload_for_register("C170", row, HEADERS["C170"])
    assert inner[0] == "C170"
    assert "999" not in inner and "x" not in inner
    assert inner[1] == "1"
    assert inner[2] == "ABC"
    assert build_sped_line(inner).startswith("|C170|1|ABC|")
